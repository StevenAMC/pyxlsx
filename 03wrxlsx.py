from openpyxl import Workbook 

wb = Workbook()
ws = wb.active

ws["A1"] = "Hello"
ws["B2"] = "there"
ws["C3"] = "from"
ws["D4"] = "Py"

wb.save("02rdwrxlsx.xlsx")

# sheet.max_row 
# It will return the last row value, you can start writing the new values from there:

# max = ws.max_row
# for row, entry in enumerate(data1, start=1):
#    st.cell(row=row+max, column=1, value=entry)
# https://stackoverflow.com/questions/37182528/how-to-append-data-using-openpyxl-python-to-excel-file-from-a-specified-row