from openpyxl import load_workbook

wb = load_workbook('01shnames.xlsx')
print(wb.sheetnames)
