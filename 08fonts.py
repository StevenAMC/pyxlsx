from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl import Workbook


wb = Workbook()

# Seleccionar sheet
ws = wb['Sheet']
font = Font(name='Calibri', size=21, bold=False, italic=False,
            vertAlign=None, underline='none', strike=False, color='FFA500')

# Font
ws['A1'].font = font
# Datos
ws['A1'] = "Hello"

wb.save("08fonts.xlsx")
